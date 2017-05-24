import requests
import openpyxl as xl
import time
import logging
import sys

from ProgressBar import ProgressBar
from FindApartment import find_apartment


def init_logging():
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")

    # init the console handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.DEBUG)
    ch.setFormatter(formatter)
    root.addHandler(ch)

    # init the DEBUG (and up/more severe) logfile handler
    dfh = logging.FileHandler('logfile.log')
    dfh.setLevel(logging.INFO)
    dfh.setFormatter(formatter)
    root.addHandler(dfh)

    # init the ERROR logfile handler
    efh = logging.FileHandler('error.log')
    efh.setLevel(logging.ERROR)
    efh.setFormatter(formatter)
    root.addHandler(efh)

init_logging()

xlsx = 'Kennemerland-with-uris.xlsx'

wb = xl.load_workbook(xlsx)
ws = wb['Resultaat BAG']
number_of_rows = ws.max_row
row_counter = 0


def get_parcel_uri(params):
    # Don't flood the server
    time.sleep(0.2)

    r = requests.get('https://brk.basisregistraties.overheid.nl/api/v1/perceel', params=params)
    # Too many requests:
    if r.status_code == 429:
        time.sleep(2)  # Give it two seconds to recuperate
        logging.warning('Retrying due to too many requests')
        r = requests.get('https://brk.basisregistraties.overheid.nl/api/v1/perceel', params=params)
    # Other bad response code:
    elif not r.status_code == 200:
        # Save results so far
        wb.save(xlsx)
        logging.error('Bad request response code %s on %s: %s' % (r.status_code, params, r.text))
        raise RuntimeError('Bad request response code %s on %s: %s' % (r.status_code, params, r.text))

    response_dict = r.json()

    # Not enough results:
    if len(response_dict['_embedded']['results']) == 0:
        raise ValueError('No results for %s' % parameters)

    # Ambiguous results:
    if len(response_dict['_embedded']['results']) > 1:
        raise ValueError('Ambiguous multiple results on request: %s' % response_dict['_embedded']['results'])

    uri = response_dict['_embedded']['results'][0]['_links']['source']['href']
    logging.debug('Found %s' % uri)

    return uri


for row in ws.iter_rows(min_row=2):  # Skip first row: column name header
    row_counter += 1
    ProgressBar.update_progress(row_counter / number_of_rows)

    # If the URI has already been filled in: skip
    if not row[4].value is None:
        logging.debug('Skipping already matched row')
        continue

    if not row[6].value is None:
        logging.debug('Skipping already tried but failed row')
        continue

    # Parse particulars from cadastral designation
    if not row[0].value:
        continue  # There could be empty rows at the end of the worksheet. We need to skip over them

    # Intermediate saves, every 100 objects
    if row_counter % 100 == 0 and row_counter > 0:
        logging.debug('Saving intermediate results')
        wb.save(xlsx)

    cadastral_designation = row[0].value

    if len(cadastral_designation.split(' ')) == 1:
        cadastral_designation = cadastral_designation[:6] + ' ' + cadastral_designation[7:]

    if len(cadastral_designation.split(' ')) == 1:  # If selective replacing didn't work...
        row[6].value = 'Kon perceelgegevens niet parsen'
        logging.error('Unable to parse parcel or apartment particulars')
        continue

    parameters = {
        'kadastraleGemeentecode': cadastral_designation.split(' ')[0][0:5],  # cadastral municipality name
        'sectie': cadastral_designation.split(' ')[0][-1],  # cadastral municipality section
        'perceelnummer': int(cadastral_designation.split(' ')[1][0:5])  # parcel number
    }

    # If labeled with an 'A' for 'apartment' rather than 'G': skip
    if cadastral_designation.split(' ')[1][5:6] == 'A':
        logging.debug('Apartment right %s' % cadastral_designation)
        apartment_designation = cadastral_designation[0:13] + '0000'
        parcel_matches = find_apartment(apartment_designation)

        if len(parcel_matches) == 0:
            logging.error('Unable to find corresponding mother parcel for apartment %s' % apartment_designation)
            row[6].value = 'Geen moederperceel gevonden'
            continue
        elif len(parcel_matches) == 1:
            # Override parcel parameter for the API lookup to the 'mother parcel'
            parameters['perceelnummer'] = int(parcel_matches[0].split(' ')[1][0:5])

            # Save the found mother parcel for reference
            row[5].value = parcel_matches[0]
        else:
            # TODO: add multiple apartment mother parcel matches
            logging.debug('Multiple mother parcels for apartment %s' % apartment_designation)
            row[6].value = 'Meerdere percelen, toegevoegd aan einde lijst'

            for match in parcel_matches:
                new_row = []
                for cell in row:
                    new_row.append(cell.value)
                # Override parcel parameter for the API lookup to the 'mother parcel'
                parameters['perceelnummer'] = int(match.split(' ')[1][0:5])

                try:
                    uri = get_parcel_uri(parameters)
                except ValueError as e:
                    logging.error(e)
                    row[6].value = str(e)
                    continue

                # Save the uri and mother parcel for reference
                new_row[4] = uri
                new_row[5] = match
                ws.append(new_row)

    try:
        uri = get_parcel_uri(parameters)
    except ValueError as e:
        logging.error(e)
        row[6].value = str(e)
        continue

    row[4].value = uri

# Finish
logging.info('Saving one last time...')
wb.save(xlsx)
logging.info('Finished!')
